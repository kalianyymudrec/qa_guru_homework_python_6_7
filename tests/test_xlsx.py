from openpyxl import load_workbook
from conftest import path_resources
import os

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    workbook = load_workbook(os.path.join(path_resources, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    sheet_value = sheet.cell(row=4, column=5).value

    sheet_row = sheet.max_row

    assert sheet_row == 51
    assert sheet_value == "France"